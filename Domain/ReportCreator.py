import json
import logging
import math
import re
from datetime import datetime
from pathlib import Path

from openpyxl.formatting.rule import CellIsRule
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font
from pandas import DataFrame, concat, ExcelWriter, read_excel

from Database.DbManager import DbManager
from Domain.AssetCollection import AssetCollection
from Domain.Enums import Direction
from Domain.InfoObject import NodeInfoObject


class ReportCreator:
    def __init__(self, collection: AssetCollection, db_manager: DbManager):
        self.collection = collection
        self.db_manager = db_manager

    def create_all_reports(self, installatie_nummer: str = None):
        df_report_pov_legacy = self.start_creating_report_pov_legacy(installatie_nummer=installatie_nummer)
        try:
            excel_name = df_report_pov_legacy['legacy_drager_naampad'].iloc[0]
        except IndexError:
            now = datetime.now()
            excel_name = f'Report_unnamed_{now.strftime("%Y%m%d_%H%M%S")}'
        if excel_name is not None and '/' in excel_name:
            excel_name = excel_name.split('/')[0]
        with ExcelWriter(path=f'Reports/{excel_name}.xlsx') as writer:
            df_report_pov_legacy.to_excel(writer, sheet_name='pov_legacy', index=False)
            print('done writing report pov legacy')
            df_report_pov_toestel = self.start_creating_report_pov_toestel(installatie_nummer=installatie_nummer)
            df_report_pov_toestel.to_excel(writer, sheet_name='pov_toestel', index=False)
            print('done writing report pov toestel')
            df_report_pov_armatuur_controller = self.start_creating_report_pov_armatuur_controller(
                installatie_nummer=installatie_nummer)
            df_report_pov_armatuur_controller.to_excel(writer, sheet_name='pov_ac', index=False)
            print('done writing report pov ac')
            df_report_pov_drager = self.start_creating_report_pov_drager(installatie_nummer=installatie_nummer)
            df_report_pov_drager.to_excel(writer, sheet_name='pov_drager', index=False)
            print('done writing report pov drager')
            df_report_pov_segment_controller = self.start_creating_report_pov_segment_controller(installatie_nummer=installatie_nummer)
            df_report_pov_segment_controller.to_excel(writer, sheet_name='pov_segm_c', index=False)
            print('done writing report pov segment controller')
            df_report_pov_montagekast = self.start_creating_report_pov_montagekast(installatie_nummer=installatie_nummer)
            df_report_pov_montagekast.to_excel(writer, sheet_name='pov_montagekast', index=False)
            print('done writing report pov montagekast')
            df_report_pov_leddriver = self.start_creating_report_pov_leddriver(installatie_nummer=installatie_nummer)
            df_report_pov_leddriver.to_excel(writer, sheet_name='pov_driver', index=False)
            print('done writing report pov leddriver')

            df = self.start_creating_asset_data_drager(installatie_nummer=installatie_nummer)
            df.to_excel(writer, sheet_name='asset_data_otl_drager', index=False)
            print('done writing asset data drager')
            df = self.start_creating_asset_data_toestel(installatie_nummer=installatie_nummer)
            df.to_excel(writer, sheet_name='asset_data_toestel', index=False)
            print('done writing asset data toestel')
            df = self.start_creating_asset_data_ac(installatie_nummer=installatie_nummer)
            df.to_excel(writer, sheet_name='asset_data_ac', index=False)
            print('done writing asset data armatuur controller')
            df = self.start_creating_asset_data_segment_controller(installatie_nummer=installatie_nummer)
            df.to_excel(writer, sheet_name='asset_data_segm_c', index=False)
            print('done writing asset data segment controller')
            df = self.start_creating_asset_data_leddriver(installatie_nummer=installatie_nummer)
            df.to_excel(writer, sheet_name='asset_data_leddriver', index=False)
            print('done writing asset data leddriver')
            df = self.start_creating_asset_data_montagekast(installatie_nummer=installatie_nummer)
            df.to_excel(writer, sheet_name='asset_data_montagekast', index=False)
            print('done writing asset data montagekast')
            summary_dict = {
                'pov_toestel_alles_ok': [
                    len(df_report_pov_toestel['alles_ok']) == df_report_pov_toestel['alles_ok'].sum()],
                'pov_drager_alles_ok': [
                    len(df_report_pov_drager['alles_ok']) == df_report_pov_drager['alles_ok'].sum()],
                'pov_armatuur_controller_alles_ok': [
                    len(df_report_pov_armatuur_controller['alles_ok']) == df_report_pov_armatuur_controller[
                        'alles_ok'].sum()],
                'pov_segment_controller_alles_ok': [
                    len(df_report_pov_segment_controller['alles_ok']) == df_report_pov_segment_controller[
                        'alles_ok'].sum()],
                'pov_leddriver_alles_ok': [
                    len(df_report_pov_leddriver['alles_ok']) == df_report_pov_leddriver['alles_ok'].sum()],
                'pov_montagekast_alles_ok': [
                    len(df_report_pov_montagekast['alles_ok']) == df_report_pov_montagekast['alles_ok'].sum()],
            }
            df_summary = DataFrame(summary_dict)
            df_summary.to_excel(writer, sheet_name='Overzicht', index=False)

        sheet_dict = {
            'pov_legacy': [f'F2:V{max(len(df_report_pov_legacy) + 1, 2)}',
                           f'X2:X{max(len(df_report_pov_legacy) + 1, 2)}',
                           f'Z2:Z{max(len(df_report_pov_legacy) + 1, 2)}'],
            'pov_toestel': [f'E2:S{max(len(df_report_pov_toestel)+1, 2)}'],
            'pov_ac': [f'E2:H{max(len(df_report_pov_armatuur_controller) + 1, 2)}',
                       f'K2:N{max(len(df_report_pov_armatuur_controller) + 1, 2)}'],
            'pov_segm_c': [f'E2:K{max(len(df_report_pov_segment_controller) + 1, 2)}'],
            'pov_drager': [f'E2:J{max(len(df_report_pov_drager) + 1, 2)}'],
            'pov_montagekast': [f'E2:H{max(len(df_report_pov_montagekast) + 1, 2)}'],
            'pov_driver': [f'E2:I{max(len(df_report_pov_leddriver) + 1, 2)}'],
            'Overzicht': [f'A2:F2'],
        }

        red_fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')
        red_font = Font(size=11, color='FF0000')
        workbook = load_workbook(f'Reports/{excel_name}.xlsx')
        for sheet_name, ranges in sheet_dict.items():
            sheet = workbook[sheet_name]
            for range_str in ranges:
                sheet.conditional_formatting.add(range_string=range_str, cfRule=CellIsRule(
                    operator='equal', formula=[0], stopIfTrue=True, fill=red_fill, font=red_font))
                sheet.conditional_formatting.add(range_string=range_str, cfRule=CellIsRule(
                    operator='equal', formula=['FALSE'], stopIfTrue=True, fill=red_fill, font=red_font))

        # move summary sheet to the front
        sheets = workbook._sheets
        index_summary = len(sheets) - 1
        summary_sheet = sheets.pop(index_summary)
        sheets.insert(0, summary_sheet)

        workbook.save(f'Reports/{excel_name}.xlsx')
        print(f'done writing file {excel_name}.xlsx')

    def start_creating_asset_data_leddriver(self, installatie_nummer: str = None) -> DataFrame:
        df = DataFrame()
        all_column_names = [
            'aanlevering_id', 'aanlevering_naam', 'uuid', 'naam', 'toestand', 'geometrie', 'datumOprichtingObject',
            'maximaalVermogen', 'maximaleAanstuurstroom', 'merk', 'modelnaam', 'protocol']

        for missing_column_name in all_column_names:
            df[missing_column_name] = None

        # get all ac's
        drivers = self.collection.get_node_objects_by_types(['onderdeel#LEDDriver'])

        for driver in drivers:
            if not driver.active:
                continue

            driver_naam = driver.attr_dict.get('AIMNaamObject.naam', '')
            if installatie_nummer is not None and not driver_naam.startswith(installatie_nummer):
                continue

            deliveries = self.db_manager.get_deliveries_by_asset_uuid(asset_uuid=driver.uuid)
            if len(deliveries) == 0:
                aanlevering_naam = ''
                aanlevering_id = ''
            else:
                aanlevering_naam = '|'.join([d.referentie for d in deliveries])
                aanlevering_id = '|'.join([d.uuid_davie for d in deliveries if d.uuid_davie is not None])

            driver_uuid = driver.uuid

            toestand = driver.attr_dict.get('AIMToestand.toestand', None)
            if toestand is not None:
                toestand = toestand[67:]

            merk = driver.attr_dict.get('LEDDriver.merk', None)
            if merk is not None:
                merk = merk[69:]

            modelnaam = driver.attr_dict.get('LEDDriver.modelnaam', None)
            if modelnaam is not None:
                modelnaam = modelnaam[74:]

            protocol = driver.attr_dict.get('LEDDriver.protocol', None)
            if protocol is not None:
                protocol = protocol[73:]

            current_toestel_dict = {
                'aanlevering_id': [aanlevering_id], 'aanlevering_naam': [aanlevering_naam],
                'uuid': [driver_uuid], 'naam': [driver_naam],
                'toestand': [toestand],
                'datumOprichtingObject': [driver.attr_dict.get('AIMObject.datumOprichtingObject', None)],
                'geometrie': [driver.attr_dict.get('loc:Locatie.geometrie', None)],
                'maximaalVermogen': [driver.attr_dict.get('LEDDriver.maximaalVermogen', None)],
                'maximaleAanstuurstroom': [driver.attr_dict.get('LEDDriver.maximaleAanstuurstroom', None)],
                'merk': [merk],
                'modelnaam': [modelnaam],
                'protocol': [protocol],
            }

            df_current = DataFrame(current_toestel_dict)
            df = concat([df, df_current])

        return df.sort_values('naam')

    def start_creating_asset_data_montagekast(self, installatie_nummer: str = None) -> DataFrame:
        df = DataFrame()
        all_column_names = [
            'aanlevering_id', 'aanlevering_naam', 'uuid', 'naam', 'toestand', 'geometrie', 'datumOprichtingObject',
            'opstelhoogte', 'verfraaid', 'afmeting', 'heeftVerlichting', 'kastmateriaal', 'ipKlasse']

        for missing_column_name in all_column_names:
            df[missing_column_name] = None

        # get all kasten
        kasten = self.collection.get_node_objects_by_types(['onderdeel#Montagekast'])

        for kast in kasten:
            if not kast.active:
                continue

            kast_naam = kast.attr_dict.get('AIMNaamObject.naam', '')
            if installatie_nummer is not None and not kast_naam.startswith(installatie_nummer):
                continue

            deliveries = self.db_manager.get_deliveries_by_asset_uuid(asset_uuid=kast.uuid)
            if len(deliveries) == 0:
                aanlevering_naam = ''
                aanlevering_id = ''
            else:
                aanlevering_naam = '|'.join([d.referentie for d in deliveries])
                aanlevering_id = '|'.join([d.uuid_davie for d in deliveries if d.uuid_davie is not None])

            kast_uuid = kast.uuid

            toestand = kast.attr_dict.get('AIMToestand.toestand', None)
            if toestand is not None:
                toestand = toestand[67:]

            verfraaid = kast.attr_dict.get('Buitenkast.verfraaid', None)
            if verfraaid is not None:
                verfraaid = verfraaid[76:]

            kastmateriaal = kast.attr_dict.get('Kast.kastmateriaal', None)
            if kastmateriaal is not None:
                kastmateriaal = kastmateriaal[69:]

            ipKlasse = kast.attr_dict.get('Buitenkast.ipKlasse', None)
            if ipKlasse is not None:
                ipKlasse = ipKlasse[81:]

            current_toestel_dict = {
                'aanlevering_id': [aanlevering_id], 'aanlevering_naam': [aanlevering_naam],
                'uuid': [kast_uuid], 'naam': [kast_naam],
                'toestand': [toestand],
                'datumOprichtingObject': [kast.attr_dict.get('AIMObject.datumOprichtingObject', None)],
                'geometrie': [kast.attr_dict.get('loc:Locatie.geometrie', None)],
                'opstelhoogte': [kast.attr_dict.get('Montagekast.opstelhoogte', None)],
                'verfraaid': [verfraaid],
                'afmeting': [kast.attr_dict.get('Kast.afmeting', None)],
                'heeftVerlichting': [kast.attr_dict.get('Kast.heeftVerlichting', None)],
                'kastmateriaal': [kastmateriaal],
                'ipKlasse': [ipKlasse],
            }

            df_current = DataFrame(current_toestel_dict)
            df = concat([df, df_current])

        return df.sort_values('naam')

    def start_creating_asset_data_ac(self, installatie_nummer: str = None) -> DataFrame:
        df = DataFrame()
        all_column_names = [
            'aanlevering_id', 'aanlevering_naam', 'uuid', 'naam', 'toestand', 'geometrie', 'datumOprichtingObject',
            'merk', 'serienummer', 'modelnaam', 'firmwareversie', 'ipAdres', 'isDummydot']

        for missing_column_name in all_column_names:
            df[missing_column_name] = None

        # get all ac's
        acs = self.collection.get_node_objects_by_types(['onderdeel#Armatuurcontroller'])

        for ac in acs:
            if not ac.active:
                continue

            toestel_naam = ac.attr_dict.get('AIMNaamObject.naam', '')
            if installatie_nummer is not None and not toestel_naam.startswith(installatie_nummer):
                continue

            deliveries = self.db_manager.get_deliveries_by_asset_uuid(asset_uuid=ac.uuid)
            if len(deliveries) == 0:
                aanlevering_naam = ''
                aanlevering_id = ''
            else:
                aanlevering_naam = '|'.join([d.referentie for d in deliveries])
                aanlevering_id = '|'.join([d.uuid_davie for d in deliveries if d.uuid_davie is not None])

            toestel_uuid = ac.uuid

            toestand = ac.attr_dict.get('AIMToestand.toestand', None)
            if toestand is not None:
                toestand = toestand[67:]

            current_toestel_dict = {
                'aanlevering_id': [aanlevering_id], 'aanlevering_naam': [aanlevering_naam],
                'uuid': [toestel_uuid], 'naam': [toestel_naam],
                'toestand': [toestand],
                'datumOprichtingObject': [ac.attr_dict.get('AIMObject.datumOprichtingObject', None)],
                'geometrie': [ac.attr_dict.get('loc:Locatie.geometrie', None)],
                'merk': [ac.attr_dict.get('Armatuurcontroller.merk', None)],
                'serienummer': [ac.attr_dict.get('Armatuurcontroller.serienummer', None)],
                'modelnaam': [ac.attr_dict.get('Armatuurcontroller.modelnaam', None)],
                'firmwareversie': [ac.attr_dict.get('FirmwareObject.firmwareversie', None)],
                'ipAdres': [ac.attr_dict.get('Armatuurcontroller.ipAdres', None)],
                'isDummydot': [ac.attr_dict.get('Armatuurcontroller.isDummydot', None)]
            }

            df_current = DataFrame(current_toestel_dict)
            df = concat([df, df_current])

        return df.sort_values('naam')

    def start_creating_asset_data_segment_controller(self, installatie_nummer: str = None) -> DataFrame:
        df = DataFrame()
        all_column_names = [
            'aanlevering_id', 'aanlevering_naam', 'uuid', 'naam', 'toestand', 'geometrie', 'datumOprichtingObject',
            'beveiligingssleutel', 'merknaam', 'modelnaam', 'batchnummer', 'dNSNaam', 'firmwareversie',
            'iPAdres', 'serienummer']

        for missing_column_name in all_column_names:
            df[missing_column_name] = None

        # get all ac's
        segm_cs = self.collection.get_node_objects_by_types(['onderdeel#Segmentcontroller'])

        for segm_c in segm_cs:
            if not segm_c.active:
                continue

            toestel_naam = segm_c.attr_dict.get('AIMNaamObject.naam', '')
            if installatie_nummer is not None and not toestel_naam.startswith(installatie_nummer):
                continue

            deliveries = self.db_manager.get_deliveries_by_asset_uuid(asset_uuid=segm_c.uuid)
            if len(deliveries) == 0:
                aanlevering_naam = ''
                aanlevering_id = ''
            else:
                aanlevering_naam = '|'.join([d.referentie for d in deliveries])
                aanlevering_id = '|'.join([d.uuid_davie for d in deliveries if d.uuid_davie is not None])

            segm_c_uuid = segm_c.uuid

            toestand = segm_c.attr_dict.get('AIMToestand.toestand', None)
            if toestand is not None:
                toestand = toestand[67:]

            merknaam = segm_c.attr_dict.get('Segmentcontroller.merknaam', None)
            if merknaam is not None:
                merknaam = merknaam[77:]

            modelnaam = segm_c.attr_dict.get('Segmentcontroller.modelnaam', None)
            if modelnaam is not None:
                modelnaam = modelnaam[83:]

            beveiligingssleutel = segm_c.attr_dict.get('Segmentcontroller.beveiligingssleutel', None)
            if beveiligingssleutel is not None:
                beveiligingssleutel = beveiligingssleutel[86:]

            current_segm_c_dict = {
                'aanlevering_id': [aanlevering_id], 'aanlevering_naam': [aanlevering_naam],
                'uuid': [segm_c_uuid], 'naam': [toestel_naam],
                'toestand': [toestand],
                'datumOprichtingObject': [segm_c.attr_dict.get('AIMObject.datumOprichtingObject', None)],
                'geometrie': [segm_c.attr_dict.get('loc:Locatie.geometrie', None)],
                'beveiligingssleutel': [beveiligingssleutel],
                'merknaam': [merknaam],
                'modelnaam': [modelnaam],
                'batchnummer': [segm_c.attr_dict.get('Controller.batchnummer', None)],
                'dNSNaam': [segm_c.attr_dict.get('Controller.dNSNaam', None)],
                'firmwareversie': [segm_c.attr_dict.get('Controller.firmwareversie', None)],
                'iPAdres': [segm_c.attr_dict.get('Controller.iPAdres', None)],
                'serienummer': [segm_c.attr_dict.get('Controller.serienummer', None)]
            }

            df_current = DataFrame(current_segm_c_dict)
            df = concat([df, df_current])

        return df.sort_values('naam')

    def start_creating_asset_data_toestel(self, installatie_nummer: str = None) -> DataFrame:
        df = DataFrame()
        all_column_names = [
            'aanlevering_id', 'aanlevering_naam', 'uuid', 'naam', 'toestand', 'geometrie', 'datumOprichtingObject',
            'modelnaam', 'kleurTemperatuur', 'merk', 'lumenOutput', 'protector', 'lichtpuntHoogte', 'lichtkleur',
            'systeemvermogen', 'besturingsconnector', 'verlichtGebied', 'aantalTeVerlichtenRijstroken', 'overhang',
            'heeftAntiVandalisme', 'verlichtingsNiveau', 'theoretischeLevensduur', 'stroomkringnummer',
            'heeftAansluitkastGeintegreerd', 'isFaunavriendelijk', 'kleurArmatuur', 'tussenafstandLED',
            'isLijnvormig']

        for missing_column_name in all_column_names:
            df[missing_column_name] = None

        # get all toestellen
        toestellen = self.collection.get_node_objects_by_types(['onderdeel#VerlichtingstoestelLED'])

        for toestel in toestellen:
            if not toestel.active:
                continue

            toestel_naam = toestel.attr_dict.get('AIMNaamObject.naam', '')
            if installatie_nummer is not None and not toestel_naam.startswith(installatie_nummer):
                continue

            deliveries = self.db_manager.get_deliveries_by_asset_uuid(asset_uuid=toestel.uuid)
            if len(deliveries) == 0:
                aanlevering_naam = ''
                aanlevering_id = ''
            else:
                aanlevering_naam = '|'.join([d.referentie for d in deliveries])
                aanlevering_id = '|'.join([d.uuid_davie for d in deliveries if d.uuid_davie is not None])

            toestel_uuid = toestel.uuid
            toestand = toestel.attr_dict.get('AIMToestand.toestand', None)
            if toestand is not None:
                toestand = toestand[67:]
            modelnaam = toestel.attr_dict.get('Verlichtingstoestel.modelnaam', None)
            if modelnaam is not None:
                modelnaam = modelnaam[84:]
            kleurTemperatuur = toestel.attr_dict.get('VerlichtingstoestelLED.kleurTemperatuur', None)
            if kleurTemperatuur is not None:
                kleurTemperatuur = kleurTemperatuur[70:]
            merk = toestel.attr_dict.get('Verlichtingstoestel.merk', None)
            if merk is not None:
                merk = merk[79:]
            lumenOutput = toestel.attr_dict.get('VerlichtingstoestelLED.lumenOutput', None)
            if lumenOutput is not None:
                lumenOutput = lumenOutput[67:]
            protector = toestel.attr_dict.get('VerlichtingstoestelLED.protector', None)
            if protector is not None:
                protector = protector[70:]
            lichtpuntHoogte = toestel.attr_dict.get('VerlichtingstoestelLED.lichtpuntHoogte', None)
            if lichtpuntHoogte is not None:
                lichtpuntHoogte = lichtpuntHoogte[76:]
            lichtkleur = toestel.attr_dict.get('VerlichtingstoestelLED.lichtkleur', None)
            if lichtkleur is not None:
                lichtkleur = lichtkleur[71:]
            besturingsconnector = toestel.attr_dict.get('VerlichtingstoestelConnector.besturingsconnector', None)
            if besturingsconnector is not None:
                besturingsconnector = besturingsconnector[103:]
            verlichtGebied = toestel.attr_dict.get('Verlichtingstoestel.verlichtGebied', None)
            if verlichtGebied is not None:
                verlichtGebied = verlichtGebied[89:]
            aantalTeVerlichtenRijstroken = toestel.attr_dict.get('VerlichtingstoestelLED.aantalTeVerlichtenRijstroken', None)
            if aantalTeVerlichtenRijstroken is not None:
                aantalTeVerlichtenRijstroken = aantalTeVerlichtenRijstroken[89:]
            overhang = toestel.attr_dict.get('VerlichtingstoestelLED.overhang', None)
            if overhang is not None:
                overhang = overhang[69:]
            verlichtingsNiveau = toestel.attr_dict.get('VerlichtingstoestelLED.verlichtingsNiveau', None)
            if verlichtingsNiveau is not None:
                verlichtingsNiveau = verlichtingsNiveau[71:]
            kleurArmatuur = toestel.attr_dict.get('VerlichtingstoestelLED.kleurArmatuur', None)
            if kleurArmatuur is not None:
                kleurArmatuur = kleurArmatuur[69:]

            current_toestel_dict = {
                'aanlevering_id': [aanlevering_id], 'aanlevering_naam': [aanlevering_naam],
                'uuid': [toestel_uuid], 'naam': [toestel_naam],
                'toestand': [toestand],
                'datumOprichtingObject': [toestel.attr_dict.get('AIMObject.datumOprichtingObject', None)],
                'geometrie': [toestel.attr_dict.get('loc:Locatie.geometrie', None)],
                'modelnaam': [modelnaam],
                'kleurTemperatuur': [kleurTemperatuur],
                'merk': [merk],
                'lumenOutput': [lumenOutput],
                'protector': [protector],
                'lichtpuntHoogte': [lichtpuntHoogte],
                'lichtkleur': [lichtkleur],
                'systeemvermogen': [toestel.attr_dict.get('Verlichtingstoestel.systeemvermogen', None)],
                'besturingsconnector': [besturingsconnector],
                'verlichtGebied': [verlichtGebied],
                'aantalTeVerlichtenRijstroken': [aantalTeVerlichtenRijstroken],
                'overhang': [overhang],
                'heeftAntiVandalisme': [toestel.attr_dict.get('VerlichtingstoestelLED.heeftAntiVandalisme', None)],
                'verlichtingsNiveau': [verlichtingsNiveau],
                'theoretischeLevensduur': [toestel.attr_dict.get('AIMObject.theoretischeLevensduur', None)],
                'stroomkringnummer': [toestel.attr_dict.get('Verlichtingstoestel.stroomkringnummer', None)],
                'heeftAansluitkastGeintegreerd': [
                    toestel.attr_dict.get('Verlichtingstoestel.heeftAansluitkastGeintegreerd', None)],
                'isFaunavriendelijk': [toestel.attr_dict.get('VerlichtingstoestelLED.isFaunavriendelijk', None)],
                'kleurArmatuur': [kleurArmatuur],
                'tussenafstandLED': [toestel.attr_dict.get('VerlichtingstoestelLED.tussenafstandLED', None)],
                'isLijnvormig': [toestel.attr_dict.get('VerlichtingstoestelLED.isLijnvormig', None)]
            }

            df_current = DataFrame(current_toestel_dict)
            df = concat([df, df_current])

        return df.sort_values('naam')

    def start_creating_asset_data_drager(self, installatie_nummer: str = None) -> DataFrame:
        df = DataFrame()
        all_column_names = [
            'aanlevering_id', 'aanlevering_naam', 'uuid', 'naam', 'toestand', 'geometrie', 'datumOprichtingObject',
            'aantalArmen', 'masttype', 'masthoogte', 'kleur', 'beschermlaag', 'heeftStopcontact', 'armlengte',
            'elekktrischeBeveiliging', 'dwarsdoorsnede', 'leverancier', 'bevestigingToestellen',
            'normeringBotsvriendelijk', 'heeftAntiVandalismeBeugel', 'theoretischeLevensduur']

        for missing_column_name in all_column_names:
            df[missing_column_name] = None

        # get all dragers
        dragers = self.collection.get_node_objects_by_types(['onderdeel#WVLichtmast', 'onderdeel#WVConsole'])

        for drager in dragers:
            if not drager.active:
                continue

            drager_naam = drager.attr_dict.get('AIMNaamObject.naam', '')
            if installatie_nummer is not None and not drager_naam.startswith(installatie_nummer):
                continue

            deliveries = self.db_manager.get_deliveries_by_asset_uuid(asset_uuid=drager.uuid)
            if len(deliveries) == 0:
                aanlevering_naam = ''
                aanlevering_id = ''
            else:
                aanlevering_naam = '|'.join([d.referentie for d in deliveries])
                aanlevering_id = '|'.join([d.uuid_davie for d in deliveries if d.uuid_davie is not None])

            drager_uuid = drager.uuid

            toestand = drager.attr_dict.get('AIMToestand.toestand', None)
            if toestand is not None:
                toestand = toestand[67:]
            aantalArmen = drager.attr_dict.get('WVLichtmast.aantalArmen', None)
            if aantalArmen is not None:
                aantalArmen = aantalArmen[76:]
            masttype = drager.attr_dict.get('Lichtmast.masttype', None)
            if masttype is not None:
                masttype = masttype[73:]
            masthoogte = drager.attr_dict.get('Lichtmast.masthoogte', None)
            if masthoogte is not None:
                masthoogte = masthoogte.get('DtuLichtmastMasthoogte.standaardHoogte', None)
            if masthoogte is not None:
                masthoogte = masthoogte[75:]
            beschermlaag = drager.attr_dict.get('Lichtmast.beschermlaag', None)
            if beschermlaag is not None:
                beschermlaag = beschermlaag[79:]
            armlengte = drager.attr_dict.get('WVLichtmast.armlengte', None)
            if armlengte is not None:
                armlengte = armlengte[76:]
            dwarsdoorsnede = drager.attr_dict.get('Lichtmast.dwarsdoorsnede', None)
            if dwarsdoorsnede is not None:
                dwarsdoorsnede = dwarsdoorsnede[86:]
            normeringBotsvriendelijk = drager.attr_dict.get('Lichtmast.normeringBotsvriendelijk', None)
            if normeringBotsvriendelijk is not None:
                normeringBotsvriendelijk = normeringBotsvriendelijk[78:]
            leverancier = drager.attr_dict.get('Lichtmast.leverancier', None)
            if leverancier is not None:
                leverancier = leverancier[76:]

            current_drager_dict = {
                'aanlevering_id': [aanlevering_id], 'aanlevering_naam': [aanlevering_naam],
                'uuid': [drager_uuid], 'naam': [drager_naam],
                'toestand': [toestand],
                'aantalArmen': [aantalArmen],
                'datumOprichtingObject': [drager.attr_dict.get('AIMObject.datumOprichtingObject', None)],
                'masttype': [masttype],
                'masthoogte': [masthoogte],
                'geometrie': [drager.attr_dict.get('loc:Locatie.geometrie', None)],
                'kleur': [drager.attr_dict.get('Lichtmast.kleur', None)],
                'beschermlaag': [beschermlaag],
                'heeftStopcontact': [drager.attr_dict.get('Lichtmast.heeftStopcontact', None)],
                'armlengte': [armlengte],
                'elekktrischeBeveiliging': [drager.attr_dict.get('EMDraagconstructie.elekktrischeBeveiliging', None)],
                'dwarsdoorsnede': [dwarsdoorsnede],
                'theoretischeLevensduur': [drager.attr_dict.get('AIMObject.theoretischeLevensduur', None)],
                'leverancier': [leverancier],
                'bevestigingToestellen': [drager.attr_dict.get('WVLichtmast.bevestigingToestellen', None)],
                'normeringBotsvriendelijk': [normeringBotsvriendelijk],
                'heeftAntiVandalismeBeugel': [drager.attr_dict.get('Lichtmast.heeftAntiVandalismeBeugel', None)],
            }

            df_current = DataFrame(current_drager_dict)
            df = concat([df, df_current])

        return df.sort_values('naam')

    def start_creating_report_pov_montagekast(self, installatie_nummer: str = None) -> DataFrame:
        df = DataFrame()
        all_column_names = [
            'aanlevering_id', 'aanlevering_naam', 'montagekast_uuid', 'montagekast_naam', 'alles_ok',
            'montagekast_naam_conform_conventie', 'relatie_naar_drager', 'relatie_naar_armatuur_controller']

        for missing_column_name in all_column_names:
            df[missing_column_name] = None

        # get all montagekasten
        montagekasten = self.collection.get_node_objects_by_types(['onderdeel#Montagekast'])

        for montagekast in montagekasten:
            if not montagekast.active:
                continue

            montagekast_naam = montagekast.attr_dict.get('AIMNaamObject.naam', '')
            if installatie_nummer is not None and not montagekast_naam.startswith(installatie_nummer):
                continue

            deliveries = self.db_manager.get_deliveries_by_asset_uuid(asset_uuid=montagekast.uuid)
            if len(deliveries) == 0:
                aanlevering_naam = ''
                aanlevering_id = ''
            else:
                aanlevering_naam = '|'.join([d.referentie for d in deliveries])
                aanlevering_id = '|'.join([d.uuid_davie for d in deliveries if d.uuid_davie is not None])

            current_mk_dict = {
                'aanlevering_id': [aanlevering_id], 'aanlevering_naam': [aanlevering_naam],
                'montagekast_uuid': [montagekast.uuid], 'montagekast_naam': [montagekast_naam],
            }

            record_dict = self.get_report_record_for_one_montagekast(montagekast=montagekast, record_dict=current_mk_dict)
            df_current = DataFrame(record_dict)
            df = concat([df, df_current])

        return df.sort_values('montagekast_naam')

    def start_creating_report_pov_leddriver(self, installatie_nummer: str = None) -> DataFrame:
        df = DataFrame()
        all_column_names = [
            'aanlevering_id', 'aanlevering_naam', 'driver_uuid', 'driver_naam', 'alles_ok',
            'driver_naam_conform_conventie', 'sturing_relatie_naar_toestel', 'bevestiging_relatie_naar_toestel',
            'relatie_van_armatuur_controller']

        for missing_column_name in all_column_names:
            df[missing_column_name] = None

        # get all leddrivers
        leddrivers = self.collection.get_node_objects_by_types(['onderdeel#LEDDriver'])

        for leddriver in leddrivers:
            if not leddriver.active:
                continue

            leddriver_naam = leddriver.attr_dict.get('AIMNaamObject.naam', '')
            if installatie_nummer is not None and not leddriver_naam.startswith(installatie_nummer):
                continue

            deliveries = self.db_manager.get_deliveries_by_asset_uuid(asset_uuid=leddriver.uuid)
            if len(deliveries) == 0:
                aanlevering_naam = ''
                aanlevering_id = ''
            else:
                aanlevering_naam = '|'.join([d.referentie for d in deliveries])
                aanlevering_id = '|'.join([d.uuid_davie for d in deliveries if d.uuid_davie is not None])

            current_mk_dict = {
                'aanlevering_id': [aanlevering_id], 'aanlevering_naam': [aanlevering_naam],
                'driver_uuid': [leddriver.uuid], 'driver_naam': [leddriver_naam],
            }

            record_dict = self.get_report_record_for_one_leddriver(leddriver=leddriver, record_dict=current_mk_dict)
            df_current = DataFrame(record_dict)
            df = concat([df, df_current])

        return df.sort_values('driver_naam')

    def start_creating_report_pov_drager(self, installatie_nummer: str = None) -> DataFrame:
        df = DataFrame()
        all_column_names = [
            'aanlevering_id', 'aanlevering_naam', 'drager_uuid', 'drager_naam', 'alles_ok',
            'drager_naam_conform_conventie', 'relatie_naar_toestel', 'relatie_naar_legacy_drager',
            'kleur_van_toepassing', 'kleur_ingevuld']

        for missing_column_name in all_column_names:
            df[missing_column_name] = None

        # get all dragers
        dragers = self.collection.get_node_objects_by_types(['onderdeel#WVLichtmast', 'onderdeel#WVConsole'])

        for drager in dragers:
            if not drager.active:
                continue

            drager_naam = drager.attr_dict.get('AIMNaamObject.naam', '')
            if installatie_nummer is not None and not drager_naam.startswith(installatie_nummer):
                continue

            deliveries = self.db_manager.get_deliveries_by_asset_uuid(asset_uuid=drager.uuid)
            if len(deliveries) == 0:
                aanlevering_naam = ''
                aanlevering_id = ''
            else:
                aanlevering_naam = '|'.join([d.referentie for d in deliveries])
                aanlevering_id = '|'.join([d.uuid_davie for d in deliveries if d.uuid_davie is not None])

            drager_uuid = drager.uuid

            current_ac_drager_dict = {
                'aanlevering_id': [aanlevering_id], 'aanlevering_naam': [aanlevering_naam],
                'drager_uuid': [drager_uuid], 'drager_naam': [drager_naam],
            }

            record_dict = self.get_report_record_for_one_drager(drager=drager, record_dict=current_ac_drager_dict)
            df_current = DataFrame(record_dict)
            df = concat([df, df_current])

        return df.sort_values('drager_naam')

    def get_report_record_for_one_leddriver(self, leddriver: NodeInfoObject, record_dict: dict) -> dict:
        leddriver_uuid = leddriver.uuid
        leddriver_naam = leddriver.attr_dict.get('AIMNaamObject.naam', '')

        alles_ok = True

        toestellen = list(self.collection.traverse_graph(
            start_uuid=leddriver_uuid, relation_types=['Sturing'], allowed_directions=[Direction.NONE],
            return_type='info_object', filtered_node_types=['onderdeel#VerlichtingstoestelLED']))
        record_dict['sturing_relatie_naar_toestel'] = [(len(toestellen) == 1)]
        alles_ok = record_dict['sturing_relatie_naar_toestel'][0] and alles_ok

        toestellen = list(self.collection.traverse_graph(
            start_uuid=leddriver_uuid, relation_types=['Bevestiging'], allowed_directions=[Direction.NONE],
            return_type='info_object', filtered_node_types=['onderdeel#VerlichtingstoestelLED']))
        record_dict['bevestiging_relatie_naar_toestel'] = [(len(toestellen) == 1)]
        alles_ok = record_dict['bevestiging_relatie_naar_toestel'][0] and alles_ok

        armatuur_controllers = list(self.collection.traverse_graph(
            start_uuid=leddriver_uuid, relation_types=['VoedtAangestuurd'], allowed_directions=[Direction.REVERSED],
            return_type='info_object', filtered_node_types=['onderdeel#Armatuurcontroller']))
        record_dict['relatie_van_armatuur_controller'] = [(len(armatuur_controllers) == 1)]
        alles_ok = record_dict['relatie_van_armatuur_controller'][0] and alles_ok

        if len(toestellen) == 1:
            toestel = toestellen[0]
            toestel_naam = toestel.attr_dict.get('AIMNaamObject.naam', '')
            record_dict['driver_naam_conform_conventie'] = [
                self.is_conform_name_convention_leddriver_with_reference(
                    leddriver_naam=leddriver_naam, toestel_naam=toestel_naam)]
        else:
            record_dict['driver_naam_conform_conventie'] = [
                self.is_conform_name_convention_leddriver_no_reference(leddriver_naam=leddriver_naam)]

        alles_ok = record_dict['driver_naam_conform_conventie'][0] and alles_ok
        record_dict['alles_ok'] = [alles_ok]
        return record_dict

    def get_report_record_for_one_montagekast(self, montagekast: NodeInfoObject, record_dict: dict) -> dict:
        montagekast_uuid = montagekast.uuid
        montagekast_naam = montagekast.attr_dict.get('AIMNaamObject.naam', '')

        alles_ok = True

        dragers = list(self.collection.traverse_graph(
            start_uuid=montagekast_uuid, relation_types=['Bevestiging'], allowed_directions=[Direction.NONE],
            return_type='info_object', filtered_node_types=['onderdeel#WVLichtmast']))
        record_dict['relatie_naar_drager'] = [(len(dragers) == 1)]
        alles_ok = record_dict['relatie_naar_drager'][0] and alles_ok

        armatuur_controllers = list(self.collection.traverse_graph(
            start_uuid=montagekast_uuid, relation_types=['VoedtAangestuurd'], allowed_directions=[Direction.WITH],
            return_type='info_object', filtered_node_types=['onderdeel#Armatuurcontroller']))
        record_dict['relatie_naar_armatuur_controller'] = [(len(armatuur_controllers) == 1)]
        alles_ok = record_dict['relatie_naar_armatuur_controller'][0] and alles_ok

        if len(dragers) == 1:
            drager = dragers[0]
            drager_naam = drager.attr_dict.get('AIMNaamObject.naam', '')
            record_dict['montagekast_naam_conform_conventie'] = [
                self.is_conform_name_convention_montagekast_with_reference(
                    montagekast_naam=montagekast_naam, drager_naam=drager_naam)]
        else:
            record_dict['montagekast_naam_conform_conventie'] = [
                self.is_conform_name_convention_montagekast_no_reference(montagekast_naam=montagekast_naam)]

        alles_ok = record_dict['montagekast_naam_conform_conventie'][0] and alles_ok
        record_dict['alles_ok'] = [alles_ok]
        return record_dict

    def get_report_record_for_one_drager(self, drager: NodeInfoObject, record_dict: dict) -> dict:
        drager_uuid = drager.uuid
        drager_naam = drager.attr_dict.get('AIMNaamObject.naam', '')

        record_dict['drager_naam_conform_conventie'] = [self.is_conform_name_convention_drager_no_reference(
            drager_naam=drager_naam)]
        alles_ok = record_dict['drager_naam_conform_conventie'][0]

        toestellen = list(self.collection.traverse_graph(
            start_uuid=drager_uuid, relation_types=['Bevestiging'], allowed_directions=[Direction.NONE],
            return_type='info_object', filtered_node_types=['onderdeel#VerlichtingstoestelLED']))
        record_dict['relatie_naar_toestel'] = [(len(toestellen) > 0)]
        alles_ok = record_dict['relatie_naar_toestel'][0] and alles_ok

        legacy_dragers = list(self.collection.traverse_graph(
            start_uuid=drager_uuid, relation_types=['HoortBij'], allowed_directions=[Direction.WITH],
            return_type='info_object', filtered_node_types=['lgc:installatie#VPLMast', 'lgc:installatie#VPConsole']))
        record_dict['relatie_naar_legacy_drager'] = [(len(legacy_dragers) > 0)]
        alles_ok = record_dict['relatie_naar_legacy_drager'][0] and alles_ok

        if drager.short_type == 'onderdeel#WVLichtmast':
            beschermlaag = drager.attr_dict.get('Lichtmast.beschermlaag', None)
            if beschermlaag == 'https://wegenenverkeer.data.vlaanderen.be/id/concept/KlDraagConstrBeschermlaag/gegalvaniseerd':
                record_dict['kleur_van_toepassing'] = [False]
            else:
                record_dict['kleur_van_toepassing'] = [True]

            record_dict['kleur_ingevuld'] = [(drager.attr_dict.get('Lichtmast.kleur', None) is not None)]
            if record_dict['kleur_van_toepassing'][0]:
                kleur_ok = record_dict['kleur_ingevuld'][0]
            else:
                kleur_ok = True

            alles_ok = kleur_ok and alles_ok

        record_dict['alles_ok'] = [alles_ok]
        return record_dict

    def start_creating_report_pov_segment_controller(self, installatie_nummer: str = None) -> DataFrame:
        df = DataFrame()
        all_column_names = [
            'aanlevering_id', 'aanlevering_naam', 'segc_uuid', 'segc_naam', 'alles_ok',
            'segmc_naam_conform_conventie', 'relatie_naar_armatuur_controller', 'relatie_naar_legacy_segmentcontroller',
            'serienummer', 'serienummer_ingevuld', 'serienummer_conform']
        for missing_column_name in all_column_names:
            df[missing_column_name] = None

        segment_controllers = self.collection.get_node_objects_by_types(['onderdeel#Segmentcontroller'])

        for segmc in segment_controllers:
            if not segmc.active:
                continue

            segmc_naam = segmc.attr_dict.get('AIMNaamObject.naam', '')
            if installatie_nummer is not None and not segmc_naam.startswith(installatie_nummer):
                continue

            deliveries = self.db_manager.get_deliveries_by_asset_uuid(asset_uuid=segmc.uuid)
            if len(deliveries) == 0:
                aanlevering_naam = ''
                aanlevering_id = ''
            else:
                aanlevering_naam = '|'.join([d.referentie for d in deliveries])
                aanlevering_id = '|'.join([d.uuid_davie for d in deliveries if d.uuid_davie is not None])

            segmc_uuid = segmc.uuid

            current_segmc_dict = {
                'aanlevering_id': [aanlevering_id], 'aanlevering_naam': [aanlevering_naam],
                'segc_uuid': [segmc_uuid], 'segc_naam': [segmc_naam],
            }

            record_dict = self.get_report_record_for_segment_controller(segment_controller=segmc,
                                                                             record_dict=current_segmc_dict)
            df_current = DataFrame(record_dict)
            df = concat([df, df_current])

        return df.sort_values('segc_naam')

    def start_creating_report_pov_armatuur_controller(self, installatie_nummer: str = None) -> DataFrame:
        df = DataFrame()
        all_column_names = [
            'aanlevering_id', 'aanlevering_naam', 'ac_uuid', 'ac_naam', 'alles_ok',
            'ac_naam_conform_conventie', 'relatie_naar_toestel', 'relatie_naar_segmentcontroller',
            'relatie_van_montagekast', 'relatie_naar_leddriver', 'serienummer', 'serienummer_ingevuld',
            'serienummer_conform', 'serienummer_uniek']

        for missing_column_name in all_column_names:
            df[missing_column_name] = None

        # get all armatuur controllers
        armatuur_controllers = self.collection.get_node_objects_by_types(['onderdeel#Armatuurcontroller'])

        for ac in armatuur_controllers:
            if not ac.active:
                continue

            ac_naam = ac.attr_dict.get('AIMNaamObject.naam', '')
            if installatie_nummer is not None and not ac_naam.startswith(installatie_nummer):
                continue

            deliveries = self.db_manager.get_deliveries_by_asset_uuid(asset_uuid=ac.uuid)
            if len(deliveries) == 0:
                aanlevering_naam = ''
                aanlevering_id = ''
            else:
                aanlevering_naam = '|'.join([d.referentie for d in deliveries])
                aanlevering_id = '|'.join([d.uuid_davie for d in deliveries if d.uuid_davie is not None])

            ac_uuid = ac.uuid

            current_ac_dict = {
                'aanlevering_id': [aanlevering_id], 'aanlevering_naam': [aanlevering_naam],
                'ac_uuid': [ac_uuid], 'ac_naam': [ac_naam],
            }

            record_dict = self.get_report_record_for_one_armatuur_controller(armatuur_controller=ac,
                                                                             record_dict=current_ac_dict)
            df_current = DataFrame(record_dict)
            df = concat([df, df_current])

        df['serienummer_uniek'] = df.duplicated(subset=['serienummer'], keep=False)
        df['serienummer_uniek'] = ~df['serienummer_uniek'] # reverse
        df['alles_ok'] = df['alles_ok'] & df['serienummer_uniek']
        #df = df.drop('serienummer', axis=1)

        return df.sort_values('ac_naam')

    def get_report_record_for_one_armatuur_controller(self, armatuur_controller: NodeInfoObject,
                                                      record_dict: dict) -> dict:
        ac_uuid = armatuur_controller.uuid
        ac_naam = armatuur_controller.attr_dict.get('AIMNaamObject.naam', '')

        record_dict['ac_naam_conform_conventie'] = [self.is_conform_name_convention_armatuur_controller_no_reference(
            ac_naam=ac_naam)]
        alles_ok = record_dict['ac_naam_conform_conventie'][0]

        toestellen = list(self.collection.traverse_graph(
            start_uuid=ac_uuid, relation_types=['Bevestiging'], allowed_directions=[Direction.NONE],
            return_type='info_object', filtered_node_types=['onderdeel#VerlichtingstoestelLED']))
        record_dict['relatie_naar_toestel'] = [(len(toestellen) == 1)]
        alles_ok = record_dict['relatie_naar_toestel'][0] and alles_ok

        segm_controllers = list(self.collection.traverse_graph(
            start_uuid=ac_uuid, relation_types=['Sturing'], allowed_directions=[Direction.NONE],
            return_type='info_object', filtered_node_types=['onderdeel#Segmentcontroller']))
        record_dict['relatie_naar_segmentcontroller'] = [(len(segm_controllers) == 1)]
        alles_ok = record_dict['relatie_naar_segmentcontroller'][0] and alles_ok

        montagekasten = list(self.collection.traverse_graph(
            start_uuid=ac_uuid, relation_types=['VoedtAangestuurd'], allowed_directions=[Direction.REVERSED],
            return_type='info_object', filtered_node_types=['onderdeel#Montagekast']))
        record_dict['relatie_van_montagekast'] = [(len(montagekasten) == 1)]
        # not included in check alles_ok

        leddrivers = list(self.collection.traverse_graph(
            start_uuid=ac_uuid, relation_types=['VoedtAangestuurd'], allowed_directions=[Direction.WITH],
            return_type='info_object', filtered_node_types=['onderdeel#LEDDriver']))
        record_dict['relatie_naar_leddriver'] = [(len(leddrivers) == 1)]
        # not included in check alles_ok

        serienummer = armatuur_controller.attr_dict.get('Armatuurcontroller.serienummer', None)
        record_dict['serienummer'] = [serienummer]
        record_dict['serienummer_ingevuld'] = [(serienummer is not None)]
        alles_ok = record_dict['serienummer_ingevuld'][0] and alles_ok

        # SLC-G3-2022-38999 example for pattern
        if serienummer is None:
            serienummer_conform = False
        else:
            serienummer_conform = re.match(r'SLC-G3-20[12]\d-[\d]{1,5}', serienummer) is not None
        record_dict['serienummer_conform'] = [serienummer_conform]
        alles_ok = serienummer_conform and alles_ok

        record_dict['alles_ok'] = [alles_ok]
        return record_dict

    def start_creating_report_pov_toestel(self, installatie_nummer: str = None) -> DataFrame:
        df = DataFrame()
        all_column_names = [
            'aanlevering_id', 'aanlevering_naam', 'toestel_uuid', 'toestel_naam', 'alles_ok',
            'toestel_naam_conform_conventie',
            'relatie_naar_armatuur_controller', 'relatie_naar_otl_of_legacy_drager',
            'aantalTeVerlichtenRijstroken_ingevuld', 'datumOprichtingObject_ingevuld', 'kleurTemperatuur_ingevuld',
            'lichtpuntHoogte_ingevuld', 'lumenOutput_ingevuld', 'overhang_ingevuld', 'verlichtingsNiveau_ingevuld',
            'merk_ingevuld', 'modelnaam_ingevuld', 'verlichtGebied_ingevuld', 'systeemvermogen_ingevuld']

        for missing_column_name in all_column_names:
            df[missing_column_name] = None

        # get all toestellen
        toestellen = self.collection.get_node_objects_by_types(['onderdeel#VerlichtingstoestelLED'])

        for toestel in toestellen:
            if not toestel.active:
                continue

            toestel_naam = toestel.attr_dict.get('AIMNaamObject.naam', '')
            if installatie_nummer is not None and not toestel_naam.startswith(installatie_nummer):
                continue

            deliveries = self.db_manager.get_deliveries_by_asset_uuid(asset_uuid=toestel.uuid)
            if len(deliveries) == 0:
                aanlevering_naam = ''
                aanlevering_id = ''
            else:
                aanlevering_naam = '|'.join([d.referentie for d in deliveries])
                aanlevering_id = '|'.join([d.uuid_davie for d in deliveries if d.uuid_davie is not None])

            toestel_uuid = toestel.uuid


            current_toestel_drager_dict = {
                'aanlevering_id': [aanlevering_id], 'aanlevering_naam': [aanlevering_naam],
                'toestel_uuid': [toestel_uuid], 'toestel_naam': [toestel_naam],
            }

            record_dict = self.get_report_record_for_one_toestel(toestel=toestel,
                                                                 record_dict=current_toestel_drager_dict)
            df_current = DataFrame(record_dict)
            df = concat([df, df_current])

        return df.sort_values('toestel_naam')

    def get_report_record_for_one_toestel(self, toestel: NodeInfoObject, record_dict: dict) -> dict:
        toestel_uuid = toestel.uuid
        toestel_naam = toestel.attr_dict.get('AIMNaamObject.naam', '')

        record_dict['toestel_naam_conform_conventie'] = [self.is_conform_name_convention_toestel_no_reference(
            toestel_naam=toestel_naam)]
        alles_ok = record_dict['toestel_naam_conform_conventie'][0]

        armatuur_controllers = list(self.collection.traverse_graph(
            start_uuid=toestel_uuid, relation_types=['Bevestiging'], allowed_directions=[Direction.NONE],
            return_type='info_object', filtered_node_types=['onderdeel#Armatuurcontroller']))
        record_dict['relatie_naar_armatuur_controller'] = [(len(armatuur_controllers) > 0)]
        alles_ok = record_dict['relatie_naar_armatuur_controller'][0] and alles_ok

        dragers = list(self.collection.traverse_graph(
            start_uuid=toestel_uuid, relation_types=['HoortBij', 'Bevestiging'],
            allowed_directions=[Direction.REVERSED, Direction.NONE], return_type='info_object',
            filtered_node_types=['onderdeel#WVLichtmast', 'onderdeel#WVConsole', 'lgc:installatie#VPBevestig']))
        record_dict['relatie_naar_otl_of_legacy_drager'] = [(len(dragers) > 0)]
        alles_ok = record_dict['relatie_naar_otl_of_legacy_drager'][0] and alles_ok

        record_dict['aantalTeVerlichtenRijstroken_ingevuld'] = [
            (toestel.attr_dict.get('VerlichtingstoestelLED.aantalTeVerlichtenRijstroken', None) is not None)]
        alles_ok = record_dict['aantalTeVerlichtenRijstroken_ingevuld'][0] and alles_ok

        record_dict['datumOprichtingObject_ingevuld'] = [
            (toestel.attr_dict.get('AIMObject.datumOprichtingObject', None) is not None)]
        alles_ok = record_dict['datumOprichtingObject_ingevuld'][0] and alles_ok

        record_dict['kleurTemperatuur_ingevuld'] = [
            (toestel.attr_dict.get('VerlichtingstoestelLED.kleurTemperatuur', None) is not None)]
        alles_ok = record_dict['kleurTemperatuur_ingevuld'][0] and alles_ok

        record_dict['lichtpuntHoogte_ingevuld'] = [
            (toestel.attr_dict.get('VerlichtingstoestelLED.lichtpuntHoogte', None) is not None)]
        alles_ok = record_dict['lichtpuntHoogte_ingevuld'][0] and alles_ok

        record_dict['lumenOutput_ingevuld'] = [
            (toestel.attr_dict.get('VerlichtingstoestelLED.lumenOutput', None) is not None)]
        alles_ok = record_dict['lumenOutput_ingevuld'][0] and alles_ok

        record_dict['overhang_ingevuld'] = [
            (toestel.attr_dict.get('VerlichtingstoestelLED.overhang', None) is not None)]
        alles_ok = record_dict['overhang_ingevuld'][0] and alles_ok

        record_dict['verlichtingsNiveau_ingevuld'] = [
            (toestel.attr_dict.get('VerlichtingstoestelLED.verlichtingsNiveau', None) is not None)]
        alles_ok = record_dict['verlichtingsNiveau_ingevuld'][0] and alles_ok

        record_dict['merk_ingevuld'] = [
            (toestel.attr_dict.get('Verlichtingstoestel.merk', None) is not None)]
        alles_ok = record_dict['merk_ingevuld'][0] and alles_ok

        record_dict['modelnaam_ingevuld'] = [
            (toestel.attr_dict.get('Verlichtingstoestel.modelnaam', None) is not None)]
        alles_ok = record_dict['modelnaam_ingevuld'][0] and alles_ok

        record_dict['verlichtGebied_ingevuld'] = [
            (toestel.attr_dict.get('Verlichtingstoestel.verlichtGebied', None) is not None)]
        alles_ok = record_dict['verlichtGebied_ingevuld'][0] and alles_ok

        record_dict['systeemvermogen_ingevuld'] = [
            (toestel.attr_dict.get('Verlichtingstoestel.systeemvermogen', None) is not None)]
        alles_ok = record_dict['systeemvermogen_ingevuld'][0] and alles_ok

        record_dict['alles_ok'] = [alles_ok]
        return record_dict

    def start_creating_report_pov_legacy(self, installatie_nummer: str = None) -> DataFrame:
        df = DataFrame()
        all_column_names = [
            'aanlevering_id', 'aanlevering_naam', 'legacy_drager_uuid', 'legacy_drager_type',
            'legacy_drager_naampad', 'legacy_drager_naampad_conform_conventie',
            'drager_verwacht', 'relatie_legacy_naar_drager_aanwezig',
            'drager_uuid', 'drager_type', 'drager_naam', 'drager_naam_conform_conventie',
            'relatie_drager_naar_toestel_aanwezig',
            'LED_toestel_1_uuid', 'LED_toestel_1_naam', 'LED_toestel_1_naam_conform_conventie',
            'relatie_naar_armatuur_controller_1_aanwezig',
            'armatuur_controller_1_uuid', 'armatuur_controller_1_naam', 'armatuur_controller_1_naam_conform_conventie',
            'legacy_drager_en_drager_binnen_5_meter',
            'legacy_drager_en_drager_identieke_geometrie', 'update_legacy_drager_geometrie',
            'legacy_drager_en_drager_gelijke_toestand', 'update_legacy_drager_toestand',
            'attributen_gelijk', 'update_legacy_drager_attributen',
            'LED_toestel_2_uuid', 'LED_toestel_2_naam', 'LED_toestel_2_naam_conform_conventie',
            'LED_toestel_3_uuid', 'LED_toestel_3_naam', 'LED_toestel_3_naam_conform_conventie',
            'LED_toestel_4_uuid', 'LED_toestel_4_naam', 'LED_toestel_4_naam_conform_conventie',
            'relatie_naar_armatuur_controller_2_aanwezig',
            'armatuur_controller_2_uuid', 'armatuur_controller_2_naam', 'armatuur_controller_2_naam_conform_conventie',
            'relatie_naar_armatuur_controller_3_aanwezig',
            'armatuur_controller_3_uuid', 'armatuur_controller_3_naam', 'armatuur_controller_3_naam_conform_conventie',
            'relatie_naar_armatuur_controller_4_aanwezig',
            'armatuur_controller_4_uuid', 'armatuur_controller_4_naam', 'armatuur_controller_4_naam_conform_conventie']
        for missing_column_name in all_column_names:
            df[missing_column_name] = None

        # get all legacy dragers
        dragers = self.collection.get_node_objects_by_types([
            'lgc:installatie#VPLMast', 'lgc:installatie#VPConsole', 'lgc:installatie#VPBevestig'])

        for drager in dragers:
            if not drager.active:
                continue

            legacy_drager_naampad = drager.attr_dict.get('NaampadObject.naampad', '')
            if installatie_nummer is not None and not legacy_drager_naampad.startswith(installatie_nummer):
                continue

            deliveries = self.db_manager.get_deliveries_by_asset_uuid(asset_uuid=drager.uuid)
            if len(deliveries) == 0:
                aanlevering_naam = ''
                aanlevering_id = ''
            else:
                aanlevering_naam = '|'.join([d.referentie for d in deliveries])
                aanlevering_id = '|'.join([d.uuid_davie for d in deliveries if d.uuid_davie is not None])
            legacy_drager_uuid = drager.uuid
            legacy_drager_type = drager.short_type.split('#')[-1]
            current_lgc_drager_dict = {
                'aanlevering_id': [aanlevering_id], 'aanlevering_naam': [aanlevering_naam],
                'legacy_drager_uuid': legacy_drager_uuid, 'legacy_drager_naampad': legacy_drager_naampad,
                'legacy_drager_type': legacy_drager_type
            }

            record_dict = self.get_report_record_for_one_lgc_drager(lgc_drager=drager,
                                                                    record_dict=current_lgc_drager_dict)
            df_current = DataFrame(record_dict)
            df = concat([df, df_current])

        return df.sort_values('legacy_drager_naampad')

    def get_report_record_for_one_lgc_drager(self, lgc_drager: NodeInfoObject, record_dict: dict) -> dict:
        legacy_drager_uuid = record_dict['legacy_drager_uuid']
        legacy_drager_naampad = record_dict['legacy_drager_naampad']
        legacy_drager_type = record_dict['legacy_drager_type']

        installatie_nummer = self.get_installatie_nummer_from_naampad(naampad=legacy_drager_naampad)
        lichtpunt_nummer = self.get_lichtpunt_nummer_from_naampad(naampad=legacy_drager_naampad)
        record_dict['legacy_drager_naampad_conform_conventie'] = (
            self.is_conform_name_convention_legacy_drager(
                legacy_drager_naampad=legacy_drager_naampad, installatie_nummer=installatie_nummer,
                lichtpunt_nummer=lichtpunt_nummer))

        record_dict['drager_verwacht'] = [legacy_drager_type != 'VPBevestig']

        toestellen: [str | NodeInfoObject]
        if record_dict['drager_verwacht'][0]:
            dragers = list(self.collection.traverse_graph(
                start_uuid=legacy_drager_uuid, relation_types=['HoortBij'], allowed_directions=[Direction.REVERSED],
                return_type='info_object', filtered_node_types=['onderdeel#WVLichtmast', 'onderdeel#WVConsole',
                                                                'onderdeel#PunctueleVerlichtingsmast']))
            if not dragers:
                logging.info(f"{legacy_drager_naampad} heeft geen HoortBij relatie naar een drager")
                record_dict['relatie_legacy_naar_drager_aanwezig'] = [False]
                return record_dict

            drager = dragers[0]
            drager_uuid = drager.uuid
            drager_naam = drager.attr_dict.get('AIMNaamObject.naam', '')
            record_dict['relatie_legacy_naar_drager_aanwezig'] = [True]
            record_dict['drager_uuid'] = [drager.uuid]
            record_dict['drager_type'] = [drager.short_type.split('#')[-1]]
            record_dict['drager_naam'] = [drager_naam]
            record_dict['drager_naam_conform_conventie'] = (
                self.is_conform_name_convention_drager(drager_name=drager_naam, installatie_nummer=installatie_nummer,
                                                       lichtpunt_nummer=lichtpunt_nummer))

            toestellen = list(self.collection.traverse_graph(
                start_uuid=drager_uuid, relation_types=['Bevestiging'], allowed_directions=[Direction.NONE],
                return_type='info_object', filtered_node_types=['onderdeel#VerlichtingstoestelLED']))
            if not toestellen:
                if drager_naam == '':
                    drager_naam = drager_uuid
                logging.info(
                    f"drager {drager_naam} van {legacy_drager_naampad} heeft geen relatie naar een LED toestel")
                record_dict['relatie_drager_naar_toestel_aanwezig'] = [False]
                return record_dict
        else:
            toestellen = list(self.collection.traverse_graph(
                start_uuid=legacy_drager_uuid, relation_types=['HoortBij'], allowed_directions=[Direction.REVERSED],
                return_type='info_object', filtered_node_types=['onderdeel#VerlichtingstoestelLED']))
            if not toestellen:
                logging.info(f"{legacy_drager_naampad} heeft geen HoortBij relatie naar een LED toestel")
                record_dict['relatie_drager_naar_toestel_aanwezig'] = [False]
                return record_dict

        record_dict['relatie_drager_naar_toestel_aanwezig'] = [True]
        toestellen_new = sorted(toestellen, key=lambda t: t.attr_dict.get('AIMNaamObject.naam', ''))
        armatuur_controllers = []

        for index, toestel in enumerate(toestellen_new):
            toestel_index = index + 1
            toestel_uuid = toestel.uuid
            toestel_name = toestel.attr_dict.get('AIMNaamObject.naam', '')
            record_dict[f'LED_toestel_{toestel_index}_uuid'] = [toestel_uuid]
            record_dict[f'LED_toestel_{toestel_index}_naam'] = [toestel_name]
            record_dict[
                f'LED_toestel_{toestel_index}_naam_conform_conventie'] = self.is_conform_name_convention_toestel(
                toestel_name=toestel_name, installatie_nummer=installatie_nummer, lichtpunt_nummer=lichtpunt_nummer,
                toestel_index=toestel_index)

            controllers = list(self.collection.traverse_graph(
                start_uuid=toestel_uuid, relation_types=['Bevestiging'], allowed_directions=[Direction.NONE],
                return_type='info_object', filtered_node_types=['onderdeel#Armatuurcontroller']))

            if not controllers:
                logging.info(f"toestel {toestel_index} van {legacy_drager_naampad} heeft geen relatie "
                             f"naar een armatuur controller")
                record_dict[f'relatie_naar_armatuur_controller_{toestel_index}_aanwezig'] = [False]
            else:
                controller = controllers[0]
                armatuur_controllers.append(controller)
                record_dict[f'relatie_naar_armatuur_controller_{toestel_index}_aanwezig'] = [True]
                record_dict[f'armatuur_controller_{toestel_index}_uuid'] = [controller.uuid]
                controller_name = controller.attr_dict.get('AIMNaamObject.naam', '')
                record_dict[f'armatuur_controller_{toestel_index}_naam'] = [controller_name]
                record_dict[f'armatuur_controller_{toestel_index}_naam_conform_conventie'] = (
                    self.is_conform_name_convention_armatuur_controller(
                        controller_name=controller_name, toestel_name=toestel_name))

        if drager:
            # geometry
            distance = self.distance_between_drager_and_legacy_drager(legacy_drager=lgc_drager, drager=drager)
            record_dict['legacy_drager_en_drager_binnen_5_meter'] = [distance <= 5.0]
            record_dict['legacy_drager_en_drager_identieke_geometrie'] = [0.0 < distance <= 0.01]
            record_dict['update_legacy_drager_geometrie'] = ''
            if distance > 5.0:
                x, y = self.get_drager_x_y(drager=drager)
                record_dict['update_legacy_drager_geometrie'] = [f'{x}|{y}']

            # toestand
            legacy_drager_toestand = lgc_drager.attr_dict.get('AIMToestand.toestand')
            drager_toestand = drager.attr_dict.get('AIMToestand.toestand')
            record_dict['legacy_drager_en_drager_gelijke_toestand'] = [legacy_drager_toestand == drager_toestand]
            if record_dict['legacy_drager_en_drager_gelijke_toestand'][0]:
                record_dict['update_legacy_drager_toestand'] = ''
            else:
                record_dict['update_legacy_drager_toestand'] = drager_toestand[67:]

        # bestekkoppeling
        # TODO

        # schadebeheerder
        # TODO

        # toezichter
        # TODO

        # vtc instructie
        # TODO?

        # attributen
        drager_dict = self.get_attribute_dict_from_otl_assets(drager=drager, toestellen=toestellen,
                                                              armatuur_controllers=armatuur_controllers)
        legacy_drager_dict = self.get_attribute_dict_from_legacy_drager(legacy_drager=lgc_drager)
        update_dict = self.get_update_dict(drager_dict=drager_dict, legacy_drager_dict=legacy_drager_dict)
        if update_dict == {}:
            record_dict['attributen_gelijk'] = [True]
            record_dict['update_legacy_drager_attributen'] = ['']
        else:
            record_dict['attributen_gelijk'] = [False]
            record_dict['update_legacy_drager_attributen'] = [json.dumps(update_dict, indent=4)]

        # aantal_te_verlichten_rijvakken_LED
        # aantal_verlichtingstoestellen
        # contractnummer_levering_LED
        # datum_installatie_LED
        # kleurtemperatuur_LED
        # LED_verlichting
        # lichtmast/bevestiging/console_buiten_gebruik
        # lichtpunthoogte_tov_rijweg
        # lumen_pakket_LED
        # overhang_LED
        # RAL_kleur_(VPLMAST/console/VVOP)
        # serienummer_armatuurcontroller_1
        # serienummer_armatuurcontroller_2
        # serienummer_armatuurcontroller_3
        # serienummer_armatuurcontroller_4
        # verlichtingsniveau_LED
        # verlichtingstoestel_merk_en_type
        # verlichtingstoestel_systeemvermogen
        # verlichtingstype

        return record_dict

    @classmethod
    def is_conform_name_convention_toestel(cls, toestel_name: str, installatie_nummer: str,
                                           lichtpunt_nummer: str, toestel_index: int) -> bool:
        return toestel_name == f'{installatie_nummer}.{lichtpunt_nummer}.WV{toestel_index}'

    @classmethod
    def get_installatie_nummer_from_toestel_name(cls, toestel_name: str) -> str:
        if toestel_name is None or not toestel_name or '.' not in toestel_name:
            return ''
        return toestel_name.split('.')[0]

    @classmethod
    def get_installatie_nummer_from_naampad(cls, naampad: str) -> str:
        if naampad is None or not naampad or '/' not in naampad:
            return ''
        return naampad.split('/')[0]

    @classmethod
    def get_lichtpunt_nummer_from_naampad(cls, naampad: str) -> str:
        if naampad is None or not naampad or '/' not in naampad:
            return ''
        lichtpunt_nummer = naampad.split('/')[2]
        if not lichtpunt_nummer[:1].isdigit():
            return lichtpunt_nummer
        inst_nr = re.search(r'[1-9][\d]*$', naampad.split('/')[0]).group()
        if inst_nr is None:
            return lichtpunt_nummer
        if lichtpunt_nummer.startswith(inst_nr):
            return lichtpunt_nummer[len(inst_nr):]
        return lichtpunt_nummer


    @classmethod
    def get_lichtpunt_nummer_from_toestel_name(cls, toestel_name: str) -> str:
        if toestel_name is None or not toestel_name or '.' not in toestel_name:
            return ''
        return toestel_name.split('.')[1]

    @classmethod
    def is_conform_name_convention_armatuur_controller(cls, controller_name: str, toestel_name: str) -> bool:
        if controller_name is None or not controller_name:
            return False
        if toestel_name is None or not toestel_name:
            return False
        return controller_name == f'{toestel_name}.AC1'

    @classmethod
    def is_conform_name_convention_drager(cls, drager_name: str, installatie_nummer: str,
                                          lichtpunt_nummer: str) -> bool:
        if drager_name is None or not drager_name:
            return False
        if installatie_nummer is None or not installatie_nummer:
            return False
        if lichtpunt_nummer is None or not lichtpunt_nummer:
            return False
        return drager_name == f'{installatie_nummer}.{lichtpunt_nummer}'

    @classmethod
    def is_conform_name_convention_legacy_drager(cls, legacy_drager_naampad: str, installatie_nummer: str,
                                                 lichtpunt_nummer: str) -> bool:
        if legacy_drager_naampad is None or not legacy_drager_naampad:
            return False
        if installatie_nummer is None or not installatie_nummer:
            return False
        if lichtpunt_nummer is None or not lichtpunt_nummer:
            return False
        if not re.match('^(A|C|G|WO|WW)[0-9]{4}$', installatie_nummer):
            return False
        return legacy_drager_naampad == f'{installatie_nummer}/{installatie_nummer}.WV/{lichtpunt_nummer}'

    @classmethod
    def distance_between_drager_and_legacy_drager(cls, legacy_drager: NodeInfoObject, drager: NodeInfoObject) -> float:
        if legacy_drager is None or drager is None:
            return 100.0
        legacy_puntlocatie = legacy_drager.attr_dict.get('loc:Locatie.puntlocatie')
        if legacy_puntlocatie is None:
            return 100.0
        legacy_puntgeometrie = legacy_puntlocatie.get('loc:3Dpunt.puntgeometrie')
        if legacy_puntgeometrie is None:
            return 100.0
        legacy_coords = legacy_puntgeometrie.get('loc:DtcCoord.lambert72')
        if legacy_coords is None:
            return 100.0
        legacy_x = legacy_coords.get('loc:DtcCoordLambert72.xcoordinaat')
        legacy_y = legacy_coords.get('loc:DtcCoordLambert72.ycoordinaat')
        if legacy_x is None or legacy_y is None:
            return 100.0

        drager_x, drager_y = cls.get_drager_x_y(drager=drager)
        if drager_x is None or drager_y is None:
            return 100.0

        return math.sqrt(abs(legacy_x - drager_x) ** 2 + abs(legacy_y - drager_y) ** 2)

    @classmethod
    def get_drager_x_y(cls, drager: NodeInfoObject) -> (float, float):
        drager_logs = drager.attr_dict.get('geo:Geometrie.log')
        if drager_logs is None:
            return None, None
        if len(drager_logs) == 0:
            return None, None
        log = next((log for log in drager_logs
                    if log.get('geo:DtcLog.niveau') == 'https://geo.data.wegenenverkeer.be/id/concept/KlLogNiveau/0'),
                   None)
        if log is None:
            log = next((log for log in drager_logs
                        if
                        log.get('geo:DtcLog.niveau') == 'https://geo.data.wegenenverkeer.be/id/concept/KlLogNiveau/-1'),
                       None)
        if log is None:
            return None, None
        drager_geometrie = log.get('geo:DtcLog.geometrie')
        if drager_geometrie is None:
            return None, None
        drager_puntgeometrie = drager_geometrie.get('geo:DtuGeometrie.punt')
        if drager_puntgeometrie is None:
            return None, None
        # use regex to get coordinates out of wkt string in drager_puntgeometrie
        drager_coords = re.match(r'POINT Z \((\d+.\d+) (\d+.\d+) (\d+)\)', drager_puntgeometrie)
        if len(drager_coords.groups()) != 3:
            return None, None
        return float(drager_coords[1]), float(drager_coords[2])

    @classmethod
    def get_attribute_dict_from_legacy_drager(cls, legacy_drager: NodeInfoObject) -> dict:
        d = {
            'aantal_te_verlichten_rijvakken_LED': legacy_drager.attr_dict.get(
                'lgc:EMObject.aantalTeVerlichtenRijvakkenLed'),
            'aantal_verlichtingstoestellen': legacy_drager.attr_dict.get('lgc:EMObject.aantalVerlichtingstoestellen'),
            'contractnummer_levering_LED': legacy_drager.attr_dict.get('lgc:EMObject.contractnummerLeveringLed'),
            'datum_installatie_LED': legacy_drager.attr_dict.get('lgc:EMObject.datumInstallatieLed'),
            'kleurtemperatuur_LED': legacy_drager.attr_dict.get('lgc:EMObject.kleurtemperatuurLed'),
            'lamp_type': legacy_drager.attr_dict.get('lgc:EMObject.lampType'),
            'LED_verlichting': legacy_drager.attr_dict.get('lgc:EMObject.ledVerlichting'),
            'lichtpunthoogte_tov_rijweg': legacy_drager.attr_dict.get('lgc:EMObject.lichtpunthoogteTovRijweg'),
            'lumen_pakket_LED': legacy_drager.attr_dict.get('lgc:EMObject.lumenPakketLed'),
            'overhang_LED': legacy_drager.attr_dict.get('lgc:EMObject.overhangLed'),
            'verlichtingsniveau_LED': legacy_drager.attr_dict.get('lgc:EMObject.verlichtingsniveauLed'),
            'verlichtingstoestel_merk_en_type': legacy_drager.attr_dict.get(
                'lgc:EMObject.verlichtingstoestelMerkEnType'),
            'verlichtingstoestel_systeemvermogen': legacy_drager.attr_dict.get(
                'lgc:EMObject.verlichtingstoestelSysteemvermogen'),
            'verlichtingstype': legacy_drager.attr_dict.get('lgc:EMObject.verlichtingstype')
        }
        if legacy_drager.short_type == 'lgc:installatie#VPLMast':
            d['drager_buiten_gebruik'] = legacy_drager.attr_dict.get('lgc:VPLMast.lichtmastBuitenGebruik')
            d['RAL_kleur'] = legacy_drager.attr_dict.get('lgc:VPLMast.ralKleurVplmast')
            d['aantal_armen'] = legacy_drager.attr_dict.get('lgc:VPLMast.aantalArmen')
            d['armlengte'] = legacy_drager.attr_dict.get('lgc:VPLMast.armlengte')
            d['paalhoogte'] = legacy_drager.attr_dict.get('lgc:VPLMast.paalhoogte')
            d['serienummer_armatuurcontroller_1'] = legacy_drager.attr_dict.get(
                'lgc:VPLMast.serienummerArmatuurcontroller1')
            d['serienummer_armatuurcontroller_2'] = legacy_drager.attr_dict.get(
                'lgc:VPLMast.serienummerArmatuurcontroller2')
            d['serienummer_armatuurcontroller_3'] = legacy_drager.attr_dict.get(
                'lgc:VPLMast.serienummerArmatuurcontroller3')
            d['serienummer_armatuurcontroller_4'] = legacy_drager.attr_dict.get(
                'lgc:VPLMast.serienummerArmatuurcontroller4')
            d['merk_en_type_armatuurcontroller_1'] = legacy_drager.attr_dict.get(
                'lgc:VPLMast.merkEnTypeArmatuurcontroller1')
            d['merk_en_type_armatuurcontroller_2'] = legacy_drager.attr_dict.get(
                'lgc:VPLMast.merkEnTypeArmatuurcontroller2')
            d['merk_en_type_armatuurcontroller_3'] = legacy_drager.attr_dict.get(
                'lgc:VPLMast.merkEnTypeArmatuurcontroller3')
            d['merk_en_type_armatuurcontroller_4'] = legacy_drager.attr_dict.get(
                'lgc:VPLMast.merkEnTypeArmatuurcontroller4')
        elif legacy_drager.short_type == 'lgc:installatie#VPConsole':
            d['drager_buiten_gebruik'] = legacy_drager.attr_dict.get('lgc:VPConsole.consoleBuitenGebruik')
            d['RAL_kleur'] = legacy_drager.attr_dict.get('lgc:VPConsole.ralKleurVpconsole')  # TODO RAL VVOP?
            d['serienummer_armatuurcontroller_1'] = legacy_drager.attr_dict.get(
                'lgc:EMObject.serienummerArmatuurcontroller')
        elif legacy_drager.short_type == 'lgc:installatie#VPBevestig':
            d['drager_buiten_gebruik'] = legacy_drager.attr_dict.get('lgc:VPBevestig.bevestigingBuitenGebruik')
            d['serienummer_armatuurcontroller_1'] = legacy_drager.attr_dict.get(
                'lgc:EMObject.serienummerArmatuurcontroller')
        return d

    @classmethod
    def get_attribute_dict_from_otl_assets(cls, drager: NodeInfoObject, toestellen: [NodeInfoObject],
                                           armatuur_controllers: [NodeInfoObject]) -> dict:
        toestel = cls.get_toestel_by_index(toestellen=toestellen, index=1)
        if toestel is None:
            return {'error': 'toestel 1 kon niet worden gevonden'}

        aantal_te_verlichten_rijvakken = toestel.attr_dict.get('VerlichtingstoestelLED.aantalTeVerlichtenRijstroken',
                                                               None)
        if aantal_te_verlichten_rijvakken is not None:
            aantal_te_verlichten_rijvakken = 'R' + aantal_te_verlichten_rijvakken[89:]

        datum_installatie_LED = toestel.attr_dict.get('AIMObject.datumOprichtingObject', None)

        kleurtemperatuur_LED = toestel.attr_dict.get('VerlichtingstoestelLED.kleurTemperatuur', None)
        if kleurtemperatuur_LED is not None:
            kleurtemperatuur_LED = 'K' + kleurtemperatuur_LED[70:]

        lichtpunthoogte = toestel.attr_dict.get('VerlichtingstoestelLED.lichtpuntHoogte')
        if lichtpunthoogte is not None:
            lichtpunthoogte = lichtpunthoogte[76:].replace('-', '.')
            try:
                lichtpunthoogte = float(lichtpunthoogte)
            except:
                logging.error(f'could not convert {lichtpunthoogte}')

        lumen_pakket_LED = toestel.attr_dict.get('VerlichtingstoestelLED.lumenOutput')
        if lumen_pakket_LED is not None:
            lumen_pakket_LED = int(lumen_pakket_LED[67:])

        overhang_LED = toestel.attr_dict.get('VerlichtingstoestelLED.overhang')
        if overhang_LED is not None:
            overhang_LED = cls.map_overhang(overhang_LED)

        verlichtingsniveau = toestel.attr_dict.get('VerlichtingstoestelLED.verlichtingsNiveau')
        if verlichtingsniveau is not None:
            verlichtingsniveau = verlichtingsniveau[71:].upper()

        merk = toestel.attr_dict.get('Verlichtingstoestel.merk')
        modelnaam = toestel.attr_dict.get('Verlichtingstoestel.modelnaam')
        merk_en_type = None
        if merk is not None and modelnaam is not None:
            merk = merk[79:].title()
            modelnaam = modelnaam[84:].title()
            if modelnaam == 'Lumi-Street':
                modelnaam = 'Lumistreet'
            merk_en_type = f'{merk} {modelnaam}'

        verlichtingstype = cls.get_verlichtingstype(toestellen)

        d = {
            'aantal_te_verlichten_rijvakken_LED': aantal_te_verlichten_rijvakken,
            'datum_installatie_LED': datum_installatie_LED,
            'kleurtemperatuur_LED': kleurtemperatuur_LED,
            'lamp_type': 'LED',
            'LED_verlichting': True,
            'lichtpunthoogte_tov_rijweg': lichtpunthoogte,
            'lumen_pakket_LED': lumen_pakket_LED,
            'overhang_LED': overhang_LED,
            'verlichtingsniveau_LED': verlichtingsniveau,
            'verlichtingstoestel_merk_en_type': merk_en_type,
            'verlichtingstoestel_systeemvermogen': toestel.attr_dict.get('Verlichtingstoestel.systeemvermogen'),
            'verlichtingstype': verlichtingstype,
            'aantal_verlichtingstoestellen': len(toestellen),

        }
        if drager.short_type == 'onderdeel#WVLichtmast':
            aantal_armen = drager.attr_dict.get('WVLichtmast.aantalArmen')
            if aantal_armen is not None:
                aantal_armen = aantal_armen[76:]
            d['aantal_armen'] = aantal_armen

            armlengte = drager.attr_dict.get('WVLichtmast.armlengte')
            if armlengte is not None:
                armlengte = armlengte[76:].replace('.', ',')
            if armlengte == 'niet-van-toepassing':
                armlengte = 'niet van toepassing'
            d['armlengte'] = [armlengte]

            paalhoogte = drager.attr_dict.get('Lichtmast.masthoogte')
            if paalhoogte is not None:
                standaard_hoogte = paalhoogte.get('DtuLichtmastMasthoogte.standaardHoogte')
                if standaard_hoogte is None:
                    raise NotImplementedError(f'{drager.attr_dict.get('AIMNaamObject.naam', '')} '
                                              f'heeft geen standaardhoogte')
                standaard_hoogte = standaard_hoogte[75:].replace('.', ',')
                d['paalhoogte'] = standaard_hoogte
            else:
                d['paalhoogte'] = None

            d['RAL_kleur'] = drager.attr_dict.get('Lichtmast.kleur')
            for index in range(1, len(toestellen) + 1):
                ac = cls.get_armatuur_controller_by_index(armatuur_controllers=armatuur_controllers, index=index)
                if ac is None:
                    continue
                d[f'serienummer_armatuurcontroller_{index}'] = ac.attr_dict.get('Armatuurcontroller.serienummer', None)
                d[f'merk_en_type_armatuurcontroller_{index}'] = cls.get_merk_en_model_ac(ac)
        return d

        # TODO
        d = {
            'contractnummer_levering_LED': drager.attr_dict.get('lgc:EMObject.contractnummerLeveringLed'),
            # via aanleveringsbestek
        }
        if drager.short_type == 'lgc:installatie#VPLMast':
            d['drager_buiten_gebruik'] = drager.attr_dict.get('lgc:VPLMast.lichtmastBuitenGebruik')

        return d

    @classmethod
    def get_update_dict(cls, drager_dict: dict, legacy_drager_dict: dict) -> dict:
        # TODO compare legacy with otl drager and create a dict that's the difference of those two
        update_dict = {}
        for key, value in drager_dict.items():
            if key not in legacy_drager_dict and value is not None:
                update_dict[key] = value
            elif value != legacy_drager_dict[key] and value is not None:
                update_dict[key] = value
        return update_dict

    @classmethod
    def map_overhang(cls, overhang_LED: str) -> str | None:
        map_dict = {
            '0-2': '0',
            '0-5': '+0.5',
            '0-5-2': '-0.5',
            '1-0': '+1.0',
            '1-0-2': '-1.0',
            '1-5': '+1.5',
            '1-5-2': '-1.5',
            '2-0': '+2.0',
            '2-0-2': '-2.0',
            '2-5': '+2.5',
            '2-5-2': '-2.5',
            '3-0': '+3.0',
            '3-0-2': '-3.0',
            '3-5': '+3.5',
            '3-5-2': '-3.5',
            '4': '-4',
            '4-2': '+4',
            '4-5': '+4.5',
            '4-5-2': '-4.5',
            '5-0': '+5.0',
            '5-0-2': '-5.0',
            '5-5': '+5.5',
            '5-5-2': '-5.5',
            '6-0': '+6.0',
            '6-0-2': '-6.0',
            '6-5': '+6.5',
            '6-5-2': '-6.5',
            '7-0': '+7.0',
            '7-0-2': '-7.0'}
        overhang = map_dict.get(overhang_LED[69:])
        if overhang is None:
            return None
        if overhang == '0':
            return '0'
        if overhang.endswith('.0'):
            overhang = overhang[:-2]

        return f'O{overhang}'

    @classmethod
    def get_verlichtingstype(cls, toestellen: [NodeInfoObject]) -> str | None:
        map_dict = {
            'afrit': ('opafrit', 2),
            'bebakening': ('bebakening', 10),
            'doorlopende-straatverlichting': ('doorlopende straatverlichting', 11),
            'fietspad': ('fietspadverlichting', 12),
            'hoofdweg': ('hoofdbaan', 4),
            'kruispunt': ('kruispunt', 5),
            'monument': ('monument', 14),
            'onderdoorgang': ('onderdoorgang', 9),
            'oprit': ('opafrit', 2),
            'parking': ('parking', 13),
            'projector': ('projector', 100),
            'punctuele-verlichting': ('punctuele verlichting', 1),
            'rotonde': ('rotonde verlichting', 6),
            'tunnelverlichting': ('tunnel verlichting', 7),
            'wisselaar': ('wisselaar', 3),
            'onderdoorgang-dag': ('onderdoorgang dag', 8)
        }

        verlichtingstype_prio = 1000
        verlichtingstype = None
        for toestel in toestellen:
            verlichtingstype_node = toestel.attr_dict.get('Verlichtingstoestel.verlichtGebied')
            if verlichtingstype_node is not None:
                verlichtingstype_node = verlichtingstype_node[89:]
                verlichtingstype_tuple = map_dict[verlichtingstype_node]
                if verlichtingstype_tuple[1] < verlichtingstype_prio:
                    verlichtingstype_prio = verlichtingstype_tuple[1]
                    verlichtingstype = verlichtingstype_tuple[0]
        return verlichtingstype

    @classmethod
    def get_armatuur_controller_by_index(cls, armatuur_controllers: [NodeInfoObject], index: int
                                         ) -> NodeInfoObject | None:
        index_str = str(index)
        for ac in armatuur_controllers:
            naam = ac.attr_dict.get('AIMNaamObject.naam')
            if naam and naam.split('.')[-2][2:3] == index_str:
                return ac

    @classmethod
    def get_toestel_by_index(cls, toestellen: [NodeInfoObject], index: int) -> NodeInfoObject | None:
        index_str = str(index)
        for toestel in toestellen:
            naam = toestel.attr_dict.get('AIMNaamObject.naam')
            if naam and naam[-1] == index_str:
                return toestel

    @staticmethod
    def is_conform_name_convention_toestel_no_reference(toestel_naam: str) -> bool:
        if '.' not in toestel_naam:
            return False
        parts = toestel_naam.split('.')
        if len(parts) != 3:
            return False
        if not re.match('^(A|C|G|WO|WW)[0-9]{4}$', parts[0]):
            return False
        if parts[2][0:2] != 'WV':
            return False
        return True

    @staticmethod
    def is_conform_name_convention_armatuur_controller_no_reference(ac_naam: str) -> bool:
        if '.' not in ac_naam:
            return False
        parts = ac_naam.split('.')
        if len(parts) != 4:
            return False
        if not re.match('^(A|C|G|WO|WW)[0-9]{4}$', parts[0]):
            return False
        if parts[2][:2] != 'WV':
            return False
        if parts[3][:2] != 'AC':
            return False
        d = parts[3][2:]
        if not parts[3][2:].isdigit():
            return False
        return True

    @staticmethod
    def is_conform_name_convention_segment_controller_no_reference(segmc_naam: str) -> bool:
        if '.' not in segmc_naam:
            return False
        parts = segmc_naam.split('.')
        if len(parts) != 2:
            return False
        if not re.match('^(A|C|G|WO|WW)[0-9]{4}$', parts[0]):
            return False
        if parts[1][:2] != 'SC':
            return False
        if not parts[1][2:].isdigit():
            return False
        return True

    @staticmethod
    def is_conform_name_convention_drager_no_reference(drager_naam: str) -> bool:
        if '.' not in drager_naam:
            return False
        parts = drager_naam.split('.')
        if len(parts) != 2:
            return False
        if not re.match('^(A|C|G|WO|WW)[0-9]{4}$', parts[0]):
            return False
        return True

    @staticmethod
    def is_conform_name_convention_leddriver_no_reference(leddriver_naam: str) -> bool:
        if '.' not in leddriver_naam:
            return False
        parts = leddriver_naam.split('.')
        if len(parts) != 4:
            return False
        if parts[3] != 'LD1':
            return False
        if not re.match('^(A|C|G|WO|WW)[0-9]{4}$', parts[0]):
            return False
        return True

    @staticmethod
    def is_conform_name_convention_leddriver_with_reference(leddriver_naam: str, toestel_naam: str) -> bool:
        parts = leddriver_naam.split('.')
        if not re.match('^(A|C|G|WO|WW)[0-9]{4}$', parts[0]):
            return False
        return leddriver_naam == f'{toestel_naam}.LD1'

    @staticmethod
    def is_conform_name_convention_montagekast_no_reference(montagekast_naam: str) -> bool:
        if '.' not in montagekast_naam:
            return False
        parts = montagekast_naam.split('.')
        if len(parts) != 3:
            return False
        if parts[2] != 'MK1':
            return False
        if not re.match('^(A|C|G|WO|WW)[0-9]{4}$', parts[0]):
            return False
        return True

    @staticmethod
    def is_conform_name_convention_montagekast_with_reference(montagekast_naam: str, drager_naam: str) -> bool:
        parts = montagekast_naam.split('.')
        if not re.match('^(A|C|G|WO|WW)[0-9]{4}$', parts[0]):
            return False
        return montagekast_naam == f'{drager_naam}.MK1'

    @classmethod
    def get_merk_en_model_ac(cls, ac: NodeInfoObject) -> str | None:
        merk = ac.attr_dict.get('Armatuurcontroller.merk')
        modelnaam = ac.attr_dict.get('Armatuurcontroller.modelnaam')
        merk_en_type = None
        if merk is not None and modelnaam is not None:
            merk_en_type = f'{merk} {modelnaam}'
        return merk_en_type

    def process_report(self, report_path: Path, em_infra_client, installatie_nummer: str = None):
        df = read_excel(report_path, 'pov_legacy')

        for index, row in df.iterrows():
            uuid = row['legacy_drager_uuid']
            if installatie_nummer is not None:
                naampad = row['legacy_drager_naampad']
                if not naampad.startswith(installatie_nummer):
                    continue
            installatie = em_infra_client.get_installatie_by_id(uuid)
            print(installatie)

            if not row['legacy_drager_en_drager_gelijke_toestand']:
                print('update toestand')
                installatie_update = em_infra_client.create_installatie_update_from_installatie(installatie)
                installatie_update.toestand = row['update_legacy_drager_toestand']
                em_infra_client.put_installatie_by_id(id=uuid, changed_installatie=installatie_update)

            if not row['legacy_drager_en_drager_binnen_5_meter']:
                print('update geometrie')
                x, y = row['update_legacy_drager_geometrie'].split('|')

                locatie = em_infra_client.get_locatie_by_installatie_id(uuid)
                print(locatie)
                locatie_update = em_infra_client.create_locatie_kenmerk_update_from_locatie_kenmerk(locatie)
                locatie_update.locatie.coordinaten.x = x
                locatie_update.locatie.coordinaten.y = y

                em_infra_client.put_locatie_kenmerk_update_by_id(id=uuid, locatie_kenmerk_update=locatie_update)

            if not row['attributen_gelijk']:
                print('update attributen')
                update_dict = json.loads(row['update_legacy_drager_attributen'])

                update_eig_dto = em_infra_client.create_update_eigenschappen_from_update_dict(
                    update_dict=update_dict, short_uri='lgc:installatie#VPLMast')
                em_infra_client.patch_eigenschapwaarden(uuid, update_eig_dto)

    def get_report_record_for_segment_controller(self, segment_controller: NodeInfoObject, record_dict: dict) -> dict:
        segmc_uuid = segment_controller.uuid
        segmc_naam = segment_controller.attr_dict.get('AIMNaamObject.naam', '')

        record_dict['segmc_naam_conform_conventie'] = [self.is_conform_name_convention_segment_controller_no_reference(
            segmc_naam=segmc_naam)]
        alles_ok = record_dict['segmc_naam_conform_conventie'][0]

        legacy_segm = list(self.collection.traverse_graph(
            start_uuid=segmc_uuid, relation_types=['HoortBij'], allowed_directions=[Direction.WITH],
            return_type='info_object', filtered_node_types=['lgc:installatie#SegC']))
        record_dict['relatie_naar_legacy_segmentcontroller'] = [(len(legacy_segm) == 1)]
        alles_ok = record_dict['relatie_naar_legacy_segmentcontroller'][0] and alles_ok

        arm_controllers = list(self.collection.traverse_graph(
            start_uuid=segmc_uuid, relation_types=['Sturing'], allowed_directions=[Direction.NONE],
            return_type='info_object', filtered_node_types=['onderdeel#Armatuurcontroller']))
        record_dict['relatie_naar_armatuur_controller'] = [(len(arm_controllers) > 1)]
        alles_ok = record_dict['relatie_naar_armatuur_controller'][0] and alles_ok

        serienummer = segment_controller.attr_dict.get('Controller.serienummer', None)
        record_dict['serienummer'] = [serienummer]
        record_dict['serienummer_ingevuld'] = [(serienummer is not None)]
        alles_ok = record_dict['serienummer_ingevuld'][0] and alles_ok

        # SLC-G3-2022-999 example for pattern
        if serienummer is None:
            serienummer_conform = False
        else:
            serienummer_conform = re.match(r'APS-G3-20[12]\d-[\d]{1,3}', serienummer) is not None
        record_dict['serienummer_conform'] = [serienummer_conform]
        alles_ok = serienummer_conform and alles_ok

        record_dict['alles_ok'] = [alles_ok]
        return record_dict
